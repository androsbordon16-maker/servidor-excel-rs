from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import requests
import io
import os
from PIL import Image as PILImage, ImageOps

app = Flask(__name__)
CORS(app)

TEMPLATE_URL = os.environ.get('TEMPLATE_URL', '')

SECCION_SHEET = {
    'Planta DC':          'DC PLANTA',
    'Distribución DC':    'DIST. Y RECT.',
    'Rectificadores':     'DIST. Y RECT.',
    'Tablero rectificadores': 'TABLERO DE AC',
    'Barra de tierras':   'TABLERO DE AC',
    'Gabinete rack':      'BANCOS ',
    'Cables baterías':    'BANCOS ',
    'Temp Bat Superior':  'TEMP.BATERIAS',
    'Temp Bat Inferior':  'TEMP.BATERIAS',
    'Temp Bat Extra':     'TEMP.BATERIAS',
    'Temp Dist Superior': 'TEMP.DISTRIBUCION',
    'Temp Dist Inferior': 'TEMP.DISTRIBUCION',
    'Temp Dist Extra':    'TEMP.DISTRIBUCION',
    'Temp Rect Izq':      'TEMP.RECTIFICADORES',
    'Temp Rect Der':      'TEMP.RECTIFICADORES',
    'Temp Rect Extra':    'TEMP.RECTIFICADORES',
    'Temp Tab Superior':  'TEMP. TABLERO AC',
    'Temp Tab Inferior':  'TEMP. TABLERO AC',
    'Temp Tab Extra':     'TEMP. TABLERO AC',
}

SECCION_ZONAS = {
    'Planta DC':              [(11,26,1,5)],
    'Distribución DC':        [(11,27,1,4),(11,27,4,8),(11,27,8,11)],
    'Rectificadores':         [(33,49,1,4),(33,49,5,10),(33,49,7,11)],
    'Tablero rectificadores': [(10,27,1,4),(10,26,4,8),(10,27,8,11)],
    'Barra de tierras':       [(38,53,1,5),(38,54,6,10)],
    'Gabinete rack':          [(11,27,1,5)],
    'Cables baterías':        [(42,57,1,4),(41,58,5,8),(41,57,8,10)],
    'Temp Bat Superior':      [(10,26,1,5),(9,25,6,10)],
    'Temp Bat Inferior':      [(28,44,1,5),(28,44,6,10)],
    'Temp Bat Extra':         [(46,61,1,5),(46,61,6,10)],
    'Temp Dist Superior':     [(10,25,1,5),(10,25,6,9)],
    'Temp Dist Inferior':     [(28,43,1,4),(28,43,6,9)],
    'Temp Dist Extra':        [(46,61,1,5),(46,61,6,9)],
    'Temp Rect Izq':          [(10,25,1,5),(10,25,6,10)],
    'Temp Rect Der':          [(28,43,1,5),(28,43,6,10)],
    'Temp Rect Extra':        [(46,61,1,5)],
    'Temp Tab Superior':      [(9,25,1,4),(10,25,6,9)],
    'Temp Tab Inferior':      [(28,44,1,4),(28,44,6,9)],
    'Temp Tab Extra':         [(46,62,1,4),(46,62,6,9)],
}

HEADER_MAX_ROW = 6

def safe_write(ws, coord, value, merge_map):
    real = merge_map.get(coord, coord)
    try:
        ws[real] = value
    except Exception:
        pass

def get_merge_map(ws):
    m = {}
    for rng in ws.merged_cells.ranges:
        top = ws.cell(rng.min_row, rng.min_col).coordinate
        for row in ws.iter_rows(min_row=rng.min_row, max_row=rng.max_row,
                                 min_col=rng.min_col, max_col=rng.max_col):
            for cell in row:
                m[cell.coordinate] = top
    return m

def limpiar_fotos_contenido(ws):
    imgs_mantener = []
    for img in ws._images:
        try:
            if img.anchor._from.row < HEADER_MAX_ROW:
                imgs_mantener.append(img)
        except:
            imgs_mantener.append(img)
    ws._images = imgs_mantener

def insertar_foto(ws, url, zona):
    try:
        resp = requests.get(url, timeout=20)
        if resp.status_code != 200:
            return
        img_data = io.BytesIO(resp.content)
        pil_img = PILImage.open(img_data)
        # Corregir orientación EXIF (fotos de celular volteadas)
        pil_img = ImageOps.exif_transpose(pil_img)
        pil_img.thumbnail((800, 600), PILImage.LANCZOS)
        if pil_img.mode in ('RGBA', 'LA', 'P'):
            pil_img = pil_img.convert('RGB')
        out = io.BytesIO()
        pil_img.save(out, format='JPEG', quality=85)
        out.seek(0)
        xl_img = XLImage(out)
        r1, r2, c1, c2 = zona
        from openpyxl.utils import get_column_letter
        col_w = sum((ws.column_dimensions[get_column_letter(c)].width or 8) if get_column_letter(c) in ws.column_dimensions else 8 for c in range(c1, c2+1))
        row_h = sum((ws.row_dimensions[r].height or 15) if r in ws.row_dimensions else 15 for r in range(r1, r2+1))
        xl_img.width = int(col_w * 7.5)
        xl_img.height = int(row_h * 1.33)
        xl_img.anchor = f"{get_column_letter(c1)}{r1}"
        ws.add_image(xl_img)
    except Exception as e:
        print(f"Error foto: {e}")

def _n(val):
    if val is None or val == '': return None
    try: return float(val)
    except: return val

def escribir_codigo_rs(wb, codigo_rs):
    valor = f'RS- {codigo_rs}'
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        mm = get_merge_map(ws)
        safe_write(ws, 'H2', valor, mm)

@app.route('/generar', methods=['POST'])
def generar():
    try:
        data = request.json
        d = data.get('datos', {})
        enc = data.get('encabezado', {})
        fotos_por_seccion = data.get('fotos', {})

        resp = requests.get(TEMPLATE_URL, timeout=30)
        wb = load_workbook(io.BytesIO(resp.content))

        for sheet in wb.sheetnames:
            limpiar_fotos_contenido(wb[sheet])

        # === CÓDIGO RS EN TODAS LAS HOJAS ===
        escribir_codigo_rs(wb, enc.get('codigo_rs', ''))

        # === DC PLANTA ===
        ws = wb['DC PLANTA']
        mm = get_merge_map(ws)
        safe_write(ws, 'D3', enc.get('planta',''), mm)
        safe_write(ws, 'G3', enc.get('fecha_servicio',''), mm)
        safe_write(ws, 'D4', enc.get('sitio',''), mm)
        safe_write(ws, 'D5', enc.get('ciudad',''), mm)
        safe_write(ws, 'I5', enc.get('numero_ventana',''), mm)
        safe_write(ws, 'H12', d.get('modelo',''), mm)
        safe_write(ws, 'H13', d.get('serie',''), mm)
        safe_write(ws, 'I15', _n(d.get('rect_total')), mm)
        safe_write(ws, 'I19', _n(d.get('rect_inst')), mm)
        safe_write(ws, 'I20', _n(d.get('cap_rect')), mm)
        safe_write(ws, 'I21', '=I20*I19/54', mm)
        safe_write(ws, 'I14', '=I15*I20/I24', mm)
        safe_write(ws, 'I23', _n(d.get('carga')), mm)
        safe_write(ws, 'I24', _n(d.get('volt_op')), mm)
        safe_write(ws, 'I25', _n(d.get('volt_ig')), mm)
        safe_write(ws, 'I26', d.get('alarmas_dc','NO'), mm)
        safe_write(ws, 'I27', '=I23/I21*I22', mm)
        safe_write(ws, 'I28', d.get('cal_pos',''), mm)
        safe_write(ws, 'I29', d.get('cal_tierra',''), mm)
        safe_write(ws, 'I30', d.get('cal_barra',''), mm)
        if d.get('nota_especial'):
            safe_write(ws, 'A32', d['nota_especial'], mm)
        if d.get('notas_dc'):
            safe_write(ws, 'A56', f'NOTAS: {d["notas_dc"]}', mm)

        # === REAPRIETE — solo filas con datos en RECT o AMP ===
        filas_excel = [38, 41, 44, 47, 50]
        rect_rows = d.get('rect_rows', [])
        filas_con_datos = [row for row in rect_rows if any([
            str(row.get('rect_izq', '')).strip(),
            str(row.get('amp_izq', '')).strip(),
            str(row.get('rect_der', '')).strip(),
            str(row.get('amp_der', '')).strip(),
        ])]
        for i, row in enumerate(filas_con_datos[:5]):
            fr = filas_excel[i]
            safe_write(ws, f'A{fr}', row.get('al',''), mm)
            if row.get('tl'): safe_write(ws, f'B{fr+1}', _n(row['tl']), mm)
            safe_write(ws, f'C{fr+1}', row.get('el',''), mm)
            safe_write(ws, f'D{fr}', f'RECT.= {row.get("rect_izq","")}  AMP= {row.get("amp_izq","")}', mm)
            safe_write(ws, f'F{fr}', f'RECT.= {row.get("rect_der","")}  AMP= {row.get("amp_der","")}', mm)
            safe_write(ws, f'G{fr}', row.get('ar',''), mm)
            if row.get('tr'): safe_write(ws, f'H{fr+1}', _n(row['tr']), mm)
            safe_write(ws, f'I{fr+1}', row.get('er',''), mm)

        # === TABLERO DE AC ===
        ws3 = wb['TABLERO DE AC']
        mm3 = get_merge_map(ws3)
        t = d.get('tableros_ac', [{}])[0] if d.get('tableros_ac') else {}
        safe_write(ws3,'B29',t.get('calibre',''),mm3)
        safe_write(ws3,'B30',_n(t.get('cables')),mm3)
        safe_write(ws3,'B31',t.get('apr1','OK'),mm3)
        safe_write(ws3,'B32',t.get('apr2','OK'),mm3)
        safe_write(ws3,'B33',t.get('apr3','OK'),mm3)
        safe_write(ws3,'H29',_n(t.get('if1')),mm3)
        safe_write(ws3,'H30',_n(t.get('if2')),mm3)
        safe_write(ws3,'H31',_n(t.get('if3')),mm3)
        safe_write(ws3,'H32',_n(t.get('vf12')),mm3)
        safe_write(ws3,'H33',_n(t.get('vf13')),mm3)
        safe_write(ws3,'H34',_n(t.get('vf23')),mm3)

        # === BANCOS ===
        ws4 = wb['BANCOS ']
        mm4 = get_merge_map(ws4)
        safe_write(ws4,'H11',d.get('rack',''),mm4)
        safe_write(ws4,'H12',d.get('bat_modelo',''),mm4)
        safe_write(ws4,'H13',d.get('bat_tipo','LITIO'),mm4)
        safe_write(ws4,'H14',_n(d.get('gab_inst')),mm4)
        safe_write(ws4,'H15',d.get('bat_marca',''),mm4)
        safe_write(ws4,'H16',d.get('bat_año',''),mm4)
        safe_write(ws4,'H17',_n(d.get('cap_banco')),mm4)
        safe_write(ws4,'H18',_n(d.get('cant_break')),mm4)
        safe_write(ws4,'H19',_n(d.get('cap_break')),mm4)
        safe_write(ws4,'I22',_n(d.get('bancos_inst')),mm4)
        safe_write(ws4,'I23',_n(d.get('cap_banco_ah')),mm4)
        safe_write(ws4,'I24','=I23*I22',mm4)
        safe_write(ws4,'C28',_n(d.get('bat_cables')),mm4)
        safe_write(ws4,'C29',d.get('bat_calibre',''),mm4)
        safe_write(ws4,'D30',d.get('bat_break_val',''),mm4)
        safe_write(ws4,'D31',d.get('bat_tierra',''),mm4)
        safe_write(ws4,'D32',d.get('bat_alarma',''),mm4)
        safe_write(ws4,'H28',_n(d.get('bat_volt')),mm4)
        safe_write(ws4,'H34',_n(d.get('bat_efic')),mm4)
        if d.get('notas_bancos'):
            safe_write(ws4,'A59',f'NOTAS: {d["notas_bancos"]}',mm4)

        # === TEMP.BATERIAS ===
        ws5 = wb['TEMP.BATERIAS']
        mm5 = get_merge_map(ws5)
        for i, gab in enumerate(d.get('gabinetes', [])[:4]):
            safe_write(ws5,f'F{38+i}',gab.get('nombre',''),mm5)
            safe_write(ws5,f'G{38+i}',gab.get('tierra',''),mm5)
        safe_write(ws5,'F34',f'ALARMAS PRESENTES:{d.get("tb_alarmas","NINGUNA")}',mm5)
        if d.get('tb_notas'):
            safe_write(ws5,'F28',f'NOTAS: {d["tb_notas"]}',mm5)

        # === TEMP.DISTRIBUCION ===
        ws6 = wb['TEMP.DISTRIBUCION']
        mm6 = get_merge_map(ws6)
        dc = [('F38','G38'),('F39','G39'),('H38','I38'),('H39','I39')]
        for i, dist in enumerate(d.get('distribuciones', [])[:4]):
            safe_write(ws6,dc[i][0],dist.get('nombre',''),mm6)
            safe_write(ws6,dc[i][1],dist.get('estado',''),mm6)
        safe_write(ws6,'F34',f'ALARMAS PRESENTES:{d.get("td_alarmas","NINGUNA")}',mm6)
        if d.get('td_notas'):
            safe_write(ws6,'F28',f'NOTAS: {d["td_notas"]}',mm6)

        # === TEMP.RECTIFICADORES ===
        ws7 = wb['TEMP.RECTIFICADORES']
        mm7 = get_merge_map(ws7)
        for i, s in enumerate(d.get('shefts_izq',[])[:4]):
            safe_write(ws7,f'F{37+i}',s.get('nombre',''),mm7)
            safe_write(ws7,f'G{37+i}',s.get('estado','OK'),mm7)
        for i, s in enumerate(d.get('shefts_der',[])[:4]):
            safe_write(ws7,f'H{37+i}',s.get('nombre',''),mm7)
            safe_write(ws7,f'I{37+i}',s.get('estado','OK'),mm7)
        safe_write(ws7,'F44',d.get('tr_limpieza','OK'),mm7)
        safe_write(ws7,'F32',f'ALARMAS PRESENTES:{d.get("tr_alarmas","NINGUNA")}',mm7)
        if d.get('tr_notas'):
            safe_write(ws7,'F28',f'NOTAS: {d["tr_notas"]}',mm7)

        # === NOTAS DIST Y RECT / TEMP TABLERO AC ===
        ws2 = wb['DIST. Y RECT.']
        mm2 = get_merge_map(ws2)
        if d.get('notas_dist'):
            safe_write(ws2,'A54',f'NOTAS: {d["notas_dist"]}',mm2)
        ws8 = wb['TEMP. TABLERO AC']
        mm8 = get_merge_map(ws8)
        if d.get('notas_temp_tablero'):
            safe_write(ws8,'F28',f'NOTAS: {d["notas_temp_tablero"]}',mm8)

        # === INSERTAR FOTOS ===
        print("📸 Fotos recibidas:", fotos_por_seccion)
        for seccion_app, urls in fotos_por_seccion.items():
            sheet_name = SECCION_SHEET.get(seccion_app)
            if not sheet_name or sheet_name not in wb.sheetnames: continue
            zonas = SECCION_ZONAS.get(seccion_app, [])
            ws_foto = wb[sheet_name]
            for idx, url in enumerate(urls):
                if idx < len(zonas):
                    insertar_foto(ws_foto, url, zonas[idx])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        codigo = enc.get('codigo_rs','RS')
        planta = enc.get('planta','REPORTE').replace(' ','_').replace('"','')
        fecha = enc.get('fecha_servicio','').replace('/','').replace(' ','_')[:15]
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"RS-{codigo}_{planta}_{fecha}.xlsx")

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
